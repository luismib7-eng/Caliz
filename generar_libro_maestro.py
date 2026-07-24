#!/usr/bin/env python3
# Genera RASTROS_UNODC_Master_DB.xlsx con la estructura completa del proyecto.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import date

AZUL_MARINO = "003366"
AZUL_UNODC = "009EDB"
GRIS = "F2F5F7"
ROJO = "D0021B"
AMBAR = "F5A623"
VERDE = "1E8E5A"

F = "Arial"
blanco = Font(name=F, sz=11, bold=True, color="FFFFFF")
cuerpo = Font(name=F, sz=10)
titulo = Font(name=F, sz=12, bold=True, color=AZUL_MARINO)
nota = Font(name=F, sz=9, italic=True, color="666666")
azul_dato = Font(name=F, sz=10, color="0000FF")

fill_marino = PatternFill("solid", fgColor=AZUL_MARINO)
fill_unodc = PatternFill("solid", fgColor=AZUL_UNODC)
fill_gris = PatternFill("solid", fgColor=GRIS)
fill_amarillo = PatternFill("solid", fgColor="FFF9C4")

centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
izq = Alignment(horizontal="left", vertical="top", wrap_text=True)
borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)

FILAS = 200
LIMITE = date(2026, 7, 29)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def encabezar(ws, headers, anchos, fila=1):
    for i, (h, w) in enumerate(zip(headers, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=h)
        c.font = blanco
        c.fill = fill_marino
        c.alignment = centro
        c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[fila].height = 42
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


# ══════════ CATÁLOGOS (se crea primero: las validaciones lo referencian) ══════
LISTAS = {
    "Tipo_Rol": ["Tomador de decisiones", "Operativo"],
    "Estatus_Vetting": ["Pendiente", "Enviado Embajada", "Aprobado"],
    "Tipo_Evento": ["Mesa de Trabajo", "Capacitación"],
    "Estatus_Catering": ["No requerido", "Solicitado a UNODC", "Confirmado", "Cubierto"],
    "Estatus_Validacion": ["Recibida", "En revisión", "Validada", "Descartada"],
    "Dependencia": [
        "Coordinación General Estratégica de Seguridad (CGES)",
        "Secretaría de Seguridad del Estado de Jalisco",
        "Fiscalía del Estado de Jalisco",
        "Fiscalía Especializada en Trata de Personas",
        "Ministerio Público",
        "Policía Cibernética",
        "Perito o personal experto técnico",
        "Comisión Ejecutiva Estatal de Atención a Víctimas",
        "Comisión de Búsqueda de Personas",
        "Otra",
    ],
}

lst = wb.create_sheet("Listas_Catalogos")
encabezar(lst, list(LISTAS.keys()), [26, 22, 20, 22, 20, 46])
for col, (k, vals) in enumerate(LISTAS.items(), start=1):
    for r, v in enumerate(vals, start=2):
        c = lst.cell(row=r, column=col, value=v)
        c.font = cuerpo
        c.border = borde
lst.cell(row=14, column=1, value="Catálogos que alimentan las listas desplegables. "
         "Si agrega valores, amplíe también el rango de validación en la pestaña correspondiente.").font = nota

rango = lambda col, n: f"Listas_Catalogos!${col}$2:${col}${n + 1}"
RG = {
    "Tipo_Rol": rango("A", 2), "Estatus_Vetting": rango("B", 3),
    "Tipo_Evento": rango("C", 2), "Estatus_Catering": rango("D", 4),
    "Estatus_Validacion": rango("E", 4), "Dependencia": rango("F", 10),
}


def validar(ws, col, rango_ref, n=FILAS, fila_ini=2):
    dv = DataValidation(type="list", formula1=rango_ref, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col}{fila_ini}:{col}{fila_ini + n - 1}")


# ══════════ 1. CONTROL_METAS_Y_KPIS ══════════
kpi = wb.create_sheet("Control_Metas_y_KPIs")
encabezar(kpi, ["ID_Meta", "Meta", "Descripción", "Objetivo_Global", "Objetivo_Jalisco",
                "Real", "Avance_%", "Semáforo", "Fuente_de_Verificación"],
          [10, 34, 52, 15, 15, 10, 12, 14, 40])

METAS = [
    ("M1", "Meta 1: Programas Diseñados",
     "Programas de formación diseñados e implementados (1 por país: México/Jalisco, Guatemala, Perú)",
     3, 1, "Documentos de programa validados por UNODC"),
    ("M2", "Meta 2: Funcionarios Capacitados",
     "Funcionarios de fuerzas del orden y fiscalías capacitados en trata vinculada a estafas en línea",
     75, 25, "Listas de asistencia — pestaña Mesas_de_Trabajo_y_Capacitaciones"),
    ("M3", "Meta 3: Investigaciones TIP / OSO",
     "Casos de trata de personas vinculados a operaciones de estafa en línea investigados (5 por país)",
     15, 5, "Reporte de la Fiscalía Especializada / carpetas de investigación"),
    ("M4", "Meta 4: Redes de Formadores y Materiales",
     "Redes ToT constituidas (3) y materiales de formación desarrollados (15)",
     3, 1, "Actas de constitución de red ToT y repositorio de materiales"),
]
for r, m in enumerate(METAS, start=2):
    for col, v in [(1, m[0]), (2, m[1]), (3, m[2]), (4, m[3]), (5, m[4]), (9, m[5])]:
        c = kpi.cell(row=r, column=col, value=v)
        c.font = cuerpo
        c.alignment = izq if col in (2, 3, 9) else centro
        c.border = borde
    # Real: dato de captura manual salvo la Meta 2, que se calcula
    if r == 3:
        kpi.cell(row=r, column=6,
                 value='=SUMIF(Mesas_de_Trabajo_y_Capacitaciones!$B$2:$B$61,"Capacitación",'
                       'Mesas_de_Trabajo_y_Capacitaciones!$F$2:$F$61)').font = cuerpo
    else:
        c = kpi.cell(row=r, column=6, value=0)
        c.font = azul_dato
        c.fill = fill_amarillo
    kpi.cell(row=r, column=6).alignment = centro
    kpi.cell(row=r, column=6).border = borde
    g = kpi.cell(row=r, column=7, value=f"=IF(D{r}=0,0,F{r}/D{r})")
    g.number_format = "0%"
    g.font, g.alignment, g.border = cuerpo, centro, borde
    h = kpi.cell(row=r, column=8,
                 value=f'=IF(G{r}>=1,"CUMPLIDA",IF(G{r}>=0.5,"EN AVANCE",'
                       f'IF(G{r}>0,"INICIADA","SIN AVANCE")))')
    h.font, h.alignment, h.border = cuerpo, centro, borde
    kpi.row_dimensions[r].height = 46

kpi.conditional_formatting.add("H2:H5", CellIsRule(
    operator="equal", formula=['"CUMPLIDA"'], fill=PatternFill("solid", fgColor=VERDE),
    font=Font(name=F, color="FFFFFF")))
kpi.conditional_formatting.add("H2:H5", CellIsRule(
    operator="equal", formula=['"EN AVANCE"'], fill=PatternFill("solid", fgColor=AMBAR)))
kpi.conditional_formatting.add("H2:H5", CellIsRule(
    operator="equal", formula=['"SIN AVANCE"'], fill=PatternFill("solid", fgColor=ROJO),
    font=Font(name=F, color="FFFFFF")))

kpi.cell(row=7, column=1, value="INDICADORES OPERATIVOS — FASE ACTUAL").font = titulo
OPS = [
    ("Beneficiarios registrados para vetting", "=COUNTA(Vetting_Beneficiarios!$B$2:$B$201)", "0"),
    ("Vetting pendiente", '=COUNTIF(Vetting_Beneficiarios!$G$2:$G$201,"Pendiente")', "0"),
    ("Vetting enviado a la Embajada", '=COUNTIF(Vetting_Beneficiarios!$G$2:$G$201,"Enviado Embajada")', "0"),
    ("Vetting aprobado", '=COUNTIF(Vetting_Beneficiarios!$G$2:$G$201,"Aprobado")', "0"),
    ("Tomadores de decisiones convocados", '=COUNTIF(Vetting_Beneficiarios!$E$2:$E$201,"Tomador de decisiones")', "0"),
    ("Dependencias participantes",
     '=SUMPRODUCT((Vetting_Beneficiarios!$D$2:$D$201<>"")/'
     'COUNTIF(Vetting_Beneficiarios!$D$2:$D$201,Vetting_Beneficiarios!$D$2:$D$201&""))', "0"),
    ("Cuestionarios de diagnóstico recibidos", '=SUMPRODUCT(--(Respuestas_Formulario_Diagnostico!$B$3:$B$202<>""))', "0"),
    ("Cuestionarios validados", '=COUNTIF(Respuestas_Formulario_Diagnostico!$C$3:$C$202,"Validada")', "0"),
    ("Días restantes para la fecha límite de vetting", "=DATE(2026,7,29)-TODAY()", "0"),
    ("Mesas de trabajo y capacitaciones programadas", "=COUNTA(Mesas_de_Trabajo_y_Capacitaciones!$B$2:$B$61)", "0"),
]
for i, (etq, form, fmt) in enumerate(OPS):
    r = 8 + i
    c = kpi.cell(row=r, column=1, value=etq)
    c.font, c.alignment, c.fill = cuerpo, izq, fill_gris
    kpi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    kpi.cell(row=r, column=2).fill = fill_gris
    v = kpi.cell(row=r, column=3, value=form)
    v.font = Font(name=F, sz=11, bold=True, color=AZUL_MARINO)
    v.number_format = fmt
    v.fill = fill_gris

kpi.cell(row=19, column=1,
         value="Celdas amarillas = captura manual. Las demás se calculan solas. "
               "Meta 2 suma los asistentes de los eventos tipo Capacitación. "
               "Objetivos tomados del One Pager RASTROS proporcionado por UNODC.").font = nota
kpi.merge_cells("A19:I19")


# ══════════ 2. VETTING_BENEFICIARIOS ══════════
vet = wb.create_sheet("Vetting_Beneficiarios")
encabezar(vet, ["ID_Beneficiario", "Nombre", "Cargo", "Dependencia", "Tipo_Rol", "Correo",
                "Estatus_Vetting", "Fecha_Limite_Envio", "Observaciones"],
          [16, 30, 30, 40, 22, 30, 20, 18, 40])

PLAZAS = [
    ("Enlace institucional", 0, "Tomador de decisiones"),
    ("Titular o representante", 1, "Tomador de decisiones"),
    ("Mando con vínculo operativo", 1, "Operativo"),
    ("Titular o representante", 2, "Tomador de decisiones"),
    ("Agente del Ministerio Público", 2, "Operativo"),
    ("Titular de la Fiscalía Especializada", 3, "Tomador de decisiones"),
    ("Investigador especializado", 3, "Operativo"),
    ("Jefatura de unidad", 5, "Tomador de decisiones"),
    ("Perito en evidencia digital", 5, "Operativo"),
    ("Enlace de atención a víctimas", 7, "Tomador de decisiones"),
    ("Enlace de búsqueda", 8, "Tomador de decisiones"),
]
deps = LISTAS["Dependencia"]
for i, (cargo, dep_i, rol) in enumerate(PLAZAS):
    r = i + 2
    vet.cell(row=r, column=2, value="Por designar").font = azul_dato
    vet.cell(row=r, column=3, value=cargo).font = azul_dato
    vet.cell(row=r, column=4, value=deps[dep_i]).font = azul_dato
    vet.cell(row=r, column=5, value=rol).font = azul_dato
    vet.cell(row=r, column=7, value="Pendiente").font = azul_dato

for r in range(2, FILAS + 2):
    vet.cell(row=r, column=1, value=f'=IF(LEN(B{r})=0,"","RAS-VET-"&TEXT(ROW()-1,"000"))')
    vet.cell(row=r, column=1).font = cuerpo
    vet.cell(row=r, column=1).alignment = centro
    d = vet.cell(row=r, column=8, value=LIMITE)
    d.number_format = "DD/MM/YYYY"
    d.alignment = centro
    d.font = cuerpo
    for col in range(2, 10):
        cc = vet.cell(row=r, column=col)
        if cc.font.color is None or cc.font.color.rgb != "FF0000FF":
            cc.font = cuerpo
        cc.alignment = izq
        cc.border = borde
    vet.cell(row=r, column=6).font = Font(name=F, sz=10, color=AZUL_UNODC)

validar(vet, "D", RG["Dependencia"])
validar(vet, "E", RG["Tipo_Rol"])
validar(vet, "G", RG["Estatus_Vetting"])

vet.conditional_formatting.add(f"G2:G{FILAS + 1}", CellIsRule(
    operator="equal", formula=['"Aprobado"'], fill=PatternFill("solid", fgColor=VERDE),
    font=Font(name=F, color="FFFFFF")))
vet.conditional_formatting.add(f"G2:G{FILAS + 1}", CellIsRule(
    operator="equal", formula=['"Enviado Embajada"'], fill=PatternFill("solid", fgColor=AMBAR)))
vet.conditional_formatting.add(f"G2:G{FILAS + 1}", CellIsRule(
    operator="equal", formula=['"Pendiente"'], fill=PatternFill("solid", fgColor=ROJO),
    font=Font(name=F, color="FFFFFF")))


# ══════════ 3. RESPUESTAS_FORMULARIO_DIAGNOSTICO ══════════
SECCIONES = [
    ("Sección I. Información general (Perfil)", [
        "P01. Correo electrónico institucional", "P02. Sexo", "P03. Edad",
        "P04. Perfil principal (Jefatura o supervisión / Operativo)",
        "P05. Institución a la que pertenece", "P06. Su rol se relaciona con",
        "P07. Rol en Seguridad Pública o Fiscalía", "P08. Unidad a la que pertenece",
        "P09. Antigüedad en el cargo",
        "P10. Años de experiencia en casos de estafas en línea",
        "P11. Años de experiencia en casos de trata de personas",
        "P12. Nivel de participación en los últimos 24 meses"]),
    ("Sección II. Detección e identificación de estafas en línea", [
        "P13. Ha sido involucrado en carpeta de investigación por estafas en línea",
        "P14. Descripción de en qué consistió la estafa",
        "P15. Perfil de las víctimas de estafas en línea",
        "P16. Qué se prometió principalmente a las víctimas",
        "P17. Perfil de las personas victimarias",
        "P18. Nacionalidad de quienes ejecutan las estafas",
        "P19. Señales o indicadores que activan la sospecha de trata",
        "P20. Identificación de call centers, bunkers o inmuebles dedicados al fraude",
        "P21. Identificación del uso de inteligencia artificial",
        "P22. Casos que involucren criptomonedas o billeteras digitales"]),
    ("Sección III. Vinculación con trata de personas (TIP), explotación y otros", [
        "P23. Casos en que se OBLIGUE a realizar trabajo o actividad en línea",
        "P24. Casos en que se ENGAÑE para realizar trabajo o actividad en línea",
        "P25. Actividades más frecuentes exigidas a la víctima",
        "P26. Tipificación del delito conforme al CNPP",
        "P27. Casos de trata de personas para cometer estafas en línea",
        "P28. Modus operandi identificado",
        "P29. Perfil de las víctimas de trata",
        "P30. Forma de captación de las víctimas",
        "P31. Traslado o facilitación de movilidad",
        "P32. Quién costeó el traslado o alojamiento inicial",
        "P33. Señales de estafa cometida contra la voluntad de la persona"]),
    ("Sección IV. Marco Legal", [
        "P34. Medida en que el marco legal facilita o dificulta la investigación",
        "P35. Claridad sobre los mecanismos de coordinación institucional",
        "P36. Nivel de participación de laboratorios forenses digitales o peritos",
        "P37. Conocimiento de unidad de enlace o fuerza de tarea conjunta TIP–Ciber",
        "P38. Necesidad de reformas legales",
        "P39. Áreas de reforma legal consideradas necesarias"]),
    ("Sección V. Recolección de pruebas e investigación (Evidencia digital)", [
        "P40. Medios de prueba determinantes",
        "P41. Dificultades durante allanamientos",
        "P42. Aplicaciones de mensajería que obstaculizan la interceptación",
        "P43. Procedimiento para preservar comunicaciones en mensajería cifrada",
        "P44. Explicación del procedimiento e institución que lo generó",
        'P45. Vinculación de "facilitadores profesionales"',
        "P46. Barreras técnicas o logísticas para obtener pruebas",
        "P47. Propuestas para mejorar la cadena de custodia digital"]),
    ("Sección VI. Protección y reparación de víctimas y sobrevivientes", [
        "P48. Aplicación de protocolos de atención integral",
        "P49. Frecuencia de uso de medidas de protección",
        "P50. Frecuencia de uso de prueba anticipada",
        "P51. Víctimas obligadas o engañadas como beneficiarias de protección",
        "P52. Dificultad para garantizar la no criminalización",
        "P53. Artículo, ley o vacío legal que constituye el principal obstáculo",
        "P54. Trato inicial que recibe quien ejecuta la estafa en línea",
        "P55. Barreras para aplicar tempranamente la no criminalización",
        "P56. Obstáculos para la reparación y restitución de derechos"]),
    ("Sección VII. Capacitación y recursos", [
        "P57. Ha recibido capacitación sobre investigación de estafas en línea",
        "P58. Ha recibido capacitación sobre investigación de trata de personas",
        "P59. Temas de capacitación considerados prioritarios",
        "P60. Recursos necesarios para mejorar la respuesta institucional"]),
    ("Sección VIII. Cooperación interinstitucional y enfoque regional", [
        "P61. Evaluación de los mecanismos de cooperación internacional",
        "P62. Mecanismos de cooperación que utilizaría",
        "P63. Actores con los que resulta más urgente coordinar",
        "P64. Principal obstáculo al solicitar información a proveedores extranjeros",
        "P65. Información o mecanismos requeridos del sector privado",
        "P66. Otras barreras institucionales, normativas o de coordinación",
        "P67. Conocimiento de alguna sentencia de trata vinculada a estafas en línea"]),
]

dia = wb.create_sheet("Respuestas_Formulario_Diagnostico")
META = ["Marca_temporal", "ID_Respuesta", "Estatus_Validacion"]
c = dia.cell(row=1, column=1, value="CONTROL DE REGISTRO")
c.font, c.fill, c.alignment = blanco, PatternFill("solid", fgColor="5A6B7B"), centro
dia.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

col = 4
total_preg = 0
for i, (tit, preguntas) in enumerate(SECCIONES):
    c = dia.cell(row=1, column=col, value=tit)
    c.font, c.alignment = blanco, centro
    c.fill = fill_marino if i % 2 == 0 else fill_unodc
    dia.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(preguntas) - 1)
    col += len(preguntas)
    total_preg += len(preguntas)

for j, h in enumerate(META + [p for _, ps in SECCIONES for p in ps], start=1):
    c = dia.cell(row=2, column=j, value=h)
    c.font = Font(name=F, sz=10, bold=True, color=AZUL_MARINO)
    c.fill, c.alignment, c.border = fill_gris, centro, borde
    dia.column_dimensions[get_column_letter(j)].width = 20 if j <= 3 else 34
dia.row_dimensions[1].height = 32
dia.row_dimensions[2].height = 72
dia.freeze_panes = "D3"

for r in range(3, FILAS + 3):
    dia.cell(row=r, column=1).number_format = "DD/MM/YYYY HH:MM"
    dia.cell(row=r, column=2, value=f'=IF(LEN(A{r})=0,"","RAS-DIA-"&TEXT(ROW()-2,"000"))').font = cuerpo
validar(dia, "C", RG["Estatus_Validacion"], n=FILAS, fila_ini=3)
assert total_preg == 67, total_preg


# ══════════ 4. MESAS_DE_TRABAJO_Y_CAPACITACIONES ══════════
mes = wb.create_sheet("Mesas_de_Trabajo_y_Capacitaciones")
encabezar(mes, ["ID_Evento", "Tipo", "Fecha", "Sede_Externa", "Estatus_Catering",
                "Asistentes_Confirmados", "Acuerdos_Principales", "Link_Minuta_PDF"],
          [14, 20, 14, 34, 20, 20, 52, 32])

EVENTOS = [
    ("Mesa de Trabajo", date(2026, 8, 26), "Por definir — sede externa UNODC", "Solicitado a UNODC", None,
     "Presentación del proyecto a personal clave con nivel de mando o comunicación con operativos.", None),
    ("Mesa de Trabajo", date(2026, 8, 27), "Por definir — sede externa UNODC", "Solicitado a UNODC", None,
     "Planeación de actividades y explicación del formulario de diagnóstico.", None),
]
for i, ev in enumerate(EVENTOS):
    r = i + 2
    for col_i, v in enumerate(ev, start=2):
        if v is not None:
            cc = mes.cell(row=r, column=col_i, value=v)
            cc.font = azul_dato
            if col_i == 3:
                cc.number_format = "DD/MM/YYYY"

for r in range(2, 62):
    mes.cell(row=r, column=1, value=f'=IF(LEN(B{r})=0,"","RAS-EVT-"&TEXT(ROW()-1,"00"))')
    mes.cell(row=r, column=1).alignment = centro
    mes.cell(row=r, column=3).number_format = "DD/MM/YYYY"
    mes.cell(row=r, column=6).number_format = "0"
    for col_i in range(1, 9):
        cc = mes.cell(row=r, column=col_i)
        if cc.font.color is None or cc.font.color.rgb != "FF0000FF":
            cc.font = cuerpo
        cc.alignment = izq
        cc.border = borde
validar(mes, "B", RG["Tipo_Evento"], n=60)
validar(mes, "E", RG["Estatus_Catering"], n=60)


# ══════════ 5. MAPEO_FORMATO_VETTING_OFICIAL ══════════
mapa = wb.create_sheet("Mapeo_Formato_VETTING_Oficial")
COLS = ["Participante o Alterno", "Nombre(s)", "Apellidos", "CURP", "Sexo",
        "Fecha de Nacimiento (mm/dd/aaaa)", "Nacimiento — Ciudad o Municipio",
        "Nacimiento — Estado", "Nacimiento — País", "Último Nivel de Estudios",
        "Conocimiento del Idioma Inglés", "Correo Electrónico", "Teléfono",
        "Residencia — Ciudad o Municipio", "Residencia — Estado", "Residencia — País",
        "Dependencia / Organización", "Puesto / Ocupación", "Rango (si aplica)",
        "¿Jefe de Unidad? (Sí/No)", "CUIP", "¿Control de Confianza Aprobado? (Sí/No)",
        "Fecha de Certificación de Control de Confianza", "Años de Experiencia en Puesto Actual",
        "Nombre de Contacto de la Dependencia", "Correo de Contacto de la Dependencia",
        "ID Unidad", "Ciudad de Partida", "Unidad en Organigrama (nivel 1)",
        "Unidad en Organigrama (nivel 2)", "Autorización del Gobierno de México"]
encabezar(mapa, COLS, [24] * len(COLS))
mapa.cell(row=2, column=1,
          value="Estructura espejo de la pestaña Participants del archivo oficial "
                "Formato VETTING 2026.xlsx. Se llena solo al momento de exportar a la Embajada; "
                "el control interno vive en Vetting_Beneficiarios.").font = nota


# ══════════ 6. LEEME ══════════
leo = wb.create_sheet("LEEME_Importacion")
leo.column_dimensions["A"].width = 118
LINEAS = [
    ("RASTROS · UNODC / Gobierno del Estado de Jalisco", "t"),
    ("Estructura de la base master — libro de importación", "t"),
    ("", ""),
    ("CÓMO IMPORTAR A GOOGLE SHEETS", "s"),
    ("1. Abra el libro RASTROS_UNODC_Master_DB en Google Sheets.", ""),
    ("2. Menú Archivo ▸ Importar ▸ Subir, y seleccione este archivo.", ""),
    ("3. En 'Ubicación de importación' elija INSERTAR HOJAS NUEVAS.", ""),
    ("4. No elija 'Reemplazar hoja': eso borraría lo que ya tiene capturado.", ""),
    ("5. Al terminar, borre la Hoja 1 anterior y esta pestaña LEEME_Importacion.", ""),
    ("", ""),
    ("QUÉ SE CAPTURA A MANO", "s"),
    ("· Celdas en AZUL: datos de ejemplo, sustitúyalos por los reales.", ""),
    ("· Celdas AMARILLAS en Control_Metas_y_KPIs: avance real de las metas 1, 3 y 4.", ""),
    ("· Todo lo demás se calcula solo. No sobrescriba las columnas de fórmula.", ""),
    ("", ""),
    ("COLUMNAS QUE NO DEBE RENOMBRAR", "s"),
    ("El tablero busca los encabezados por su nombre exacto, con guion bajo y sin acentos", ""),
    ("añadidos: ID_Beneficiario, Nombre, Cargo, Dependencia, Tipo_Rol, Correo,", ""),
    ("Estatus_Vetting, Fecha_Limite_Envio, Observaciones. Si los cambia, dejará de leerlos.", ""),
    ("", ""),
    ("PESTAÑAS DEL LIBRO", "s"),
    ("Control_Metas_y_KPIs — 4 metas del proyecto e indicadores operativos calculados.", ""),
    ("Vetting_Beneficiarios — control de propuestas y su estatus ante la Embajada.", ""),
    ("Respuestas_Formulario_Diagnostico — 67 preguntas en 8 secciones.", ""),
    ("Mesas_de_Trabajo_y_Capacitaciones — logística y acuerdos de cada evento.", ""),
    ("Listas_Catalogos — alimenta las listas desplegables. No la borre.", ""),
    ("Mapeo_Formato_VETTING_Oficial — espejo de las 31 columnas del formato de la Embajada.", ""),
    ("", ""),
    ("FUENTES", "s"),
    ("Metas y objetivos: One Pager RASTROS (UNODC).", ""),
    ("67 preguntas: Formulario RASTROS_final_Jalisco.", ""),
    ("31 columnas de vetting: Formato VETTING 2026.xlsx, pestaña Participants.", ""),
    ("Fechas y acuerdos logísticos: minuta de la reunión con UNODC del 22/07/2026.", ""),
]
for i, (txt, tipo) in enumerate(LINEAS, start=1):
    c = leo.cell(row=i, column=1, value=txt)
    if tipo == "t":
        c.font = Font(name=F, sz=14, bold=True, color=AZUL_MARINO)
    elif tipo == "s":
        c.font = Font(name=F, sz=11, bold=True, color="FFFFFF")
        c.fill = fill_marino
    else:
        c.font = cuerpo
    c.alignment = izq

wb.move_sheet("LEEME_Importacion", offset=-6)
import sys
wb.save(sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/outputs/RASTROS_UNODC_Master_DB.xlsx")
print("Pestañas:", wb.sheetnames)
print("Preguntas del diagnóstico:", total_preg)
